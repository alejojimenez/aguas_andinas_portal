import os
import time
import shutil
import requests
import re
import glob
import fitz

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
from datetime import datetime,timedelta

class Scraper_Aguas():

    def __init__(self,url, email, password, driver_path):
        print(url, email, password, driver_path)
        self.url = url
        self.email = email
        self.password = password
        self.driver_path = driver_path

    def wait(self, seconds):
        return WebDriverWait(self.driver, seconds)

    def close(self):
        self.driver.close()
        self.driver = None

    def quit(self):
        self.driver.quit()
        self.driver = None   

    def dic_datos(self,mes_texto):
        dic = {'ENE':'01','FEB':'02','MAR':'03','ABR':'04','MAY':'05','JUN':'06','JUL':'07',
               'AGO':'08','SEP':'09','OCT':'10','NOV':'11','DIC':'12'}

        mes_oficial = dic[mes_texto]
        
        return mes_oficial

    def login(self):
        
        driver_exe = '.domain\\chromedriver.exe'
        credencials = '.\\config\\credenciales.xlsx'

        print('Entrando en la funcion login...')
        print('----------------------------------------------------------------------')
        
        #Seteo variables
        email = self.email
        url = self.url
        driver_path = self.driver_path
        password  = self.password
        
        options = webdriver.ChromeOptions()
            
        self.driver = webdriver.Chrome(driver_path, options=options)
        self.driver.get(url)
        self.driver.maximize_window()
        self.driver.implicitly_wait(40)

        # Controlar evento alert() Notificacion
        try:
            alert = WebDriverWait(self.driver, 35).until(EC.alert_is_present())
            alert = Alert(self.driver)
            alert.accept()
                    
        except:
            print("No se encontró ninguna alerta.")  

        #Botones ID que utilizaremos para logear
        selector_ingreso_cuenta = 'wlogin_login'
        selector_rut_input = 'rut2'
        selector_password_input = 'clave'
        selector_login_button = 'b_login_1'

        # Seleccionar campo cuenta
        intentos = 0
        reintentar = True
        while (reintentar):
            try:
                print('Try en la funcion login para el campo cuenta...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_cuenta= self.driver.find_element(By.ID, selector_ingreso_cuenta)
                element_cuenta.click()
                reintentar = False
            except:    
                print('Exception en la funcion click campo cuenta')
                print('----------------------------------------------------------------------')
                reintentar = intentos <= 3                

        time.sleep(10) 
        # Seleccionar campo rut y setear usuario
        intentos_rut = 0
        reintentar = True
        while (reintentar):
            try:
                print('Try en la funcion click para el campo rut...', intentos_rut)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_username = self.driver.find_element(By.ID, selector_rut_input)
                element_username.click()
                element_username.send_keys(email)
                reintentar = False
            except:    
                print('Exception en la funcion click rut cliente')
                print('----------------------------------------------------------------------')
                reintentar = intentos_rut <= 3
                
        # Seleccionar campo clave y setear clave
        intentos = 0
        reintentar = True
        while (reintentar):
            try:
                print('Try en la funcion click para el campo rut...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_password = self.driver.find_element(By.ID, selector_password_input)
                element_password.click()
                element_password.send_keys(password)
                reintentar = False
            except:    
                print('Exception en la funcion click rut cliente')
                print('----------------------------------------------------------------------')
                reintentar = intentos <= 3

        # Seleccionar boton y hacer en boton ingresar
        intentos = 0
        reintentar = True
        while (reintentar):
            try:
                print('Try en la funcion click_element_xpath para el boton ingresar...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_button_ingresar = self.driver.find_element(By.ID, selector_login_button)
                element_button_ingresar.click()
                reintentar = False
            except:    
                print('Exception en la funcion click_element_xpath')
                print('----------------------------------------------------------------------')
                reintentar = intentos <= 3

    def scrapping_aguas(self,posicion):

        #Buscamos la tabla que contiene el boton de boletas
        WebDriverWait(self.driver, 30).until(
        EC.presence_of_element_located((By.ID,'scta')))
        boton_sucursal= self.driver.find_element(By.ID,'scta')
        boton_sucursal.click()

        #Buscamos las sucursales y generamos lista con los nombres
        lista_sucursales = []
        boton_sucursal_abierto = self.driver.find_elements(By.ID,'selCu')
        for index,boton in enumerate(boton_sucursal_abierto):
            #print(index,'-',boton.get_attribute('innerText'))
            sucursal = boton.get_attribute('innerText')
            resultado = re.sub(r"[^a-zA-Z ]", "", sucursal).strip()
            print(resultado)
            lista_sucursales.append(resultado)
        
        cantidad_sociedades = len(lista_sucursales)
        print(f'el largo de los locales es {cantidad_sociedades}')
        
        while posicion < cantidad_sociedades+1:
            
            #Boton de las sociedades
            intento_sociedades = 0
            reintentar_sociedades = True
            while (reintentar_sociedades):
                try:
                    time.sleep(5)
                    boton_sucursal_of = self.driver.find_elements(By.ID,'selCu')
                    time.sleep(5)    
                    menu_sucursal = boton_sucursal_of[posicion]
                    menu_sucursal.click()
                    reintentar_sociedades = False
                    time.sleep(5)
                except:    
                    print('menu sociedades no esta clicleable')
                    print('----------------------------------------------------------------------')
                    self.driver.execute_script("window.scrollTo(0, 0)")
                    reintentar_sociedades = intento_sociedades <= 5

            #Boton resumen boletas
            intento_resumen= 0
            reintentar_resumen = True
            while (reintentar_resumen):
                try:
                    boton_sucursal_of = self.driver.find_element(By.ID,'pestanaResumenBoletas')
                    time.sleep(2)    
                    boton_sucursal_of.click()
                    reintentar_resumen = False
                    time.sleep(5)
                except:    
                    print('menu resumen boletas no esta clicleable')
                    print('----------------------------------------------------------------------')
                    reintentar_resumen = intento_resumen <= 3

            #Boton ver mas boletas           
            intento_ver_mas= 0
            reintentar_ver_mas = True
            while (reintentar_ver_mas):
                try:
                    boton_sucursal_of = self.driver.find_element(By.LINK_TEXT,'Ver más documentos')
                    time.sleep(2)    
                    boton_sucursal_of.click()
                    reintentar_ver_mas = False
                    time.sleep(5)
                except:    
                    print('link ver mas boletas no esta clicleable')
                    print('----------------------------------------------------------------------')
                    self.driver.execute_script("window.scrollBy(0, 200)")
                    reintentar_ver_mas = intento_ver_mas <= 5
                    
            #Encontrar la tabla y sus elementos          
            intento_ver_tabla= 0
            reintentar_ver_tabla = True
            while (reintentar_ver_tabla):
                try:
                    tabla_element = self.driver.find_element(By.XPATH,'//*[@id="myTablePer"]')
                    time.sleep(2)
                    celdas = tabla_element.find_elements(By.TAG_NAME,'td')
                    reintentar_ver_tabla = False
                    time.sleep(5)
                except:    
                    print('Buscando tabla con datos')
                    print('----------------------------------------------------------------------')
                    reintentar_ver_tabla = intento_ver_tabla <= 5

            fila = 1
            while fila < 14:            
                
                if fila == 13:
                    #Siguiente pagina
                    intentos_cambio_hoja = 0
                    reintentar_cambio = True
                    while (reintentar_cambio):
                        try:
                            tabla_element_cambio = self.driver.find_element(By.XPATH,f'//*[@id="tabs-1"]/div[1]/ul/li[3]')
                            time.sleep(2)
                            tabla_element_cambio.click()
                            reintentar_cambio = False
                        except:
                            print('no hay numero de factura')
                            print('----------------------------------------------------------------------')
                            reintentar_cambio = intentos_cambio_hoja <= 3
                    
                #Numero de factura
                intentos_factura = 0
                reintentar_factura = True
                while (reintentar_factura):
                    try:
                        tabla_element_nfact = self.driver.find_element(By.XPATH,f'//*[@id="myTablePer"]/tbody/tr[{fila}]/td[1]')
                        factura = tabla_element_nfact.text
                        time.sleep(2)
                        reintentar_factura = False
                    except:
                        print('no hay numero de factura')
                        print('----------------------------------------------------------------------')
                        reintentar_factura = intentos_factura <= 3
                
                #Mes y año
                intentos_fecha= 0
                reintentar_fecha = True
                while (reintentar_fecha):
                    try: 
                        tabla_element_fecha = self.driver.find_element(By.XPATH,f'//*[@id="myTablePer"]/tbody/tr[{fila}]/td[2]')
                        fecha = tabla_element_fecha.text
                        partes = fecha.split("/")
                        mes = str(partes[0])
                        año = str(partes[1])
                        time.sleep(2)
                        reintentar_fecha = False
                    except:
                        print('no hay fecha')
                        print('----------------------------------------------------------------------')
                        reintentar_fecha = intentos_fecha <= 3
                
                #Boton descarga
                intentos_down= 0
                reintentar_down= True                
                while (reintentar_down):
                    try:
                        
                        tabla_element_descarga = self.driver.find_element(By.XPATH,f'//*[@id="myTablePer"]/tbody/tr[{fila}]/td[7]')
                        time.sleep(2)
                        tabla_element_descarga.click()
                        reintentar_down = False
                    except:
                        print('no hay boton de descarga')
                        print('----------------------------------------------------------------------')
                        reintentar_down = intentos_down <= 3    
        
                time.sleep(5)

                # Esperar a que se abra la ventana emergente
                intentos = 0
                reintentar_ventana = True
                while (reintentar_ventana):
                    try:
                        # Obtiene el identificador de la ventana actual
                        current_window = self.driver.current_window_handle
                        print('Ventana principal: ', current_window)
                        print('----------------------------------------------------------------------')
                        print('Try en la funcion manejo de ventanas abiertas...', intentos)
                        print('----------------------------------------------------------------------')
                        intentos += 1
                        window_handles_all = self.driver.window_handles
                        reintentar_ventana = False
                            
                    except:    
                        print('Exception en la funcion manejo de ventanas abiertas')
                        print('----------------------------------------------------------------------')
                        print('----------------------------------------------------------------------')
                        time.sleep(60) #espera para que cargue ventana emergente
                                    
                        reintentar_ventana = intentos <= 3
                                    
                # Obtiene los identificadores de las ventanas abiertas
                self.driver.implicitly_wait(35)
                print('Ventanas abiertas: ', window_handles_all, len(window_handles_all))
                print('----------------------------------------------------------------------')            
                                    
                # Cambiar al manejo de la ventana emergente
                for window_handle in window_handles_all:
                    if window_handle != current_window:
                        self.driver.switch_to.window(window_handle)
                        print('Ventana emergente: ', window_handle)
                        print('----------------------------------------------------------------------')
                        break

                # Esperar hasta que el elemento esté presente en la página
                self.driver.implicitly_wait(35)
                            
                try:
                    # Obtener la URL de la ventana emergente
                    ventana_emergente_url = self.driver.current_url
                    print("URL de la ventana emergente:", ventana_emergente_url)
                    print('----------------------------------------------------------------------')

                    # Realizar una solicitud GET para obtener la data binaria del documento
                    response = requests.get(ventana_emergente_url, stream=True)

                    # Obtener el nombre del archivo a partir de los datos del proceso de descarga
                    folder_path = './input/'
                    file_name = folder_path + str(factura) +"_"+ str(lista_sucursales[posicion])+"_"+ str(mes)+"_"+str(año)+".pdf" 

                    # Guardar la data binaria en un archivo PDF
                    with open(file_name, 'wb') as file:
                        response.raw.decode_content = True
                        shutil.copyfileobj(response.raw, file)
                        print("Guardando archivo:", file_name)
                        print('----------------------------------------------------------------------')
                                    
                except:    
                    print("No se encontró el elemento con el id especificado...")
                    print('----------------------------------------------------------------------')
                                
                    count += 1
                    print('Conteo de documentos: ', count)
                    print('----------------------------------------------------------------------')                

                # Cerrar la ventana emergente
                time.sleep(5)
                self.driver.close()
                time.sleep(15)                            

                # Cambiar de nuevo al manejo de ventana principal
                self.driver.switch_to.window(current_window)
                print('Cual ventana es: ', current_window)
                print('----------------------------------------------------------------------')
                        
                time.sleep(10)
                
                print('pasamos al siguiente archivo')
                fila +=1 
            break
        print('Pasamos a la siguiente sociedad')

    def archivos(self):
        
        folder_path = './input/'
        output_path = './output/'
        
        #Revisamos si hay archivos pdf en la carpeta input
        archivos_pdf = glob.glob(os.path.join(folder_path, '*.pdf'))

        #Si no encuentra archivos es porque no se realizo la ejecucion correcta y hay que mandar mail
        if not archivos_pdf:
            print(f'No se encontraron archivos PDF en la carpeta "{folder_path}".')
        else:
            #Si encuentra me entregara todos los documentos con los que trabajaremos
            print(f'Se encontraron los siguientes archivos PDF en la carpeta "{folder_path}":')
            
            for archivo in archivos_pdf:
                
                nombre_oficial = archivo.replace('./input','')
                            
                with fitz.open(archivo) as pdf_documento:
                    texto_completo = ''

                    for pagina_num in range(pdf_documento.page_count):
                        pagina = pdf_documento.load_page(pagina_num)
                        texto_completo += pagina.get_text()
                    
                    lista_limpia = [elemento.strip() for elemento in texto_completo.split('\n')]
                    
                    #Posicion 1
                    elemento_a_buscar = 'FACTURA ELECTRÓNICA'
                    try:
                        posicion_1 = lista_limpia.index(elemento_a_buscar)
                        n_factura_bruto = lista_limpia[posicion_1+1]
                        n_factura = n_factura_bruto.replace("Nº ", "")
                    except:
                        print('elemento no se encuentra disponible')
                        n_factura = ''

                    #Posicion 2
                    texto_a_verificar = 'RUTA:'
                    posicion_2 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_2 = idx
                            rut_bruto = lista_limpia[posicion_2+1]
                            rut = rut_bruto.replace('R.U.T.: ','')
                    
                    #Posicion 3
                    elemento_a_buscar = 'VENCIMIENTO'
                    try:
                        posicion_3 = lista_limpia.index(elemento_a_buscar)
                        valor = lista_limpia[posicion_3-1]
                        if valor.startswith('GIRO:'):
                            n_cuenta = lista_limpia[posicion_3-6]
                        elif not valor.startswith('GIRO:'):
                            n_cuenta = lista_limpia[posicion_3-1]   
                    except:
                        print('elemento no se encuentra disponible')
                        n_cuenta = ''
                        
                    #Posicion 4
                    texto_a_verificar = 'GIRO:'
                    posicion_4 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_4 = idx
                            giro_bruto = lista_limpia[posicion_4]
                            giro = giro_bruto.replace('GIRO: ','')

                    #Posicion 5
                    elemento_a_buscar = 'VENCIMIENTO'
                    try:
                        posicion_5 = lista_limpia.index(elemento_a_buscar)
                        valor_evaluar = lista_limpia[posicion_5+2]
                        if valor_evaluar.startswith('$'):
                            fecha_vencimiento_bruto = lista_limpia[posicion_5+3]
                            partes = fecha_vencimiento_bruto.split('-')
                            dia_vencimiento = partes[0]
                            mes_vencimiento = self.dic_datos(partes[1])
                            año_vencimiento = partes[2]
                            fecha_vencimiento = dia_vencimiento+'-'+mes_vencimiento+'-'+año_vencimiento
                        elif not valor_evaluar.startswith('$'):
                            fecha_vencimiento_bruto = lista_limpia[posicion_5+2]
                            partes = fecha_vencimiento_bruto.split('-')
                            dia_vencimiento = partes[0]
                            mes_vencimiento = self.dic_datos(partes[1])
                            año_vencimiento = partes[2]
                            fecha_vencimiento = dia_vencimiento+'-'+mes_vencimiento+'-'+año_vencimiento       
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_vencimiento = ''
                        
                    #Posicion 7
                    elemento_a_buscar = 'CARGO FIJO'
                    try:
                        posicion_7 = lista_limpia.index(elemento_a_buscar)
                        cargo_fijo = lista_limpia[posicion_7+1]
                    except:
                        print('elemento no se encuentra disponible')
                        cargo_fijo = ''
                    
                    #Posicion 8 y 9
                    elemento_a_buscar = 'CONSUMO AGUA'
                    try:
                        posicion_8 = lista_limpia.index(elemento_a_buscar)
                        cantidad_consumo_agua = lista_limpia[posicion_8+1]
                        monto_consumo_agua_b = lista_limpia[posicion_8+2]
                        monto_consumo_agua = monto_consumo_agua_b.replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_consumo_agua = ''
                        monto_consumo_agua = ''

                    #Posicion 10 y 11
                    elemento_a_buscar = 'RECOLECCION'
                    try:
                        posicion_10 = lista_limpia.index(elemento_a_buscar)
                        cantidad_recoleccion = lista_limpia[posicion_10+1]
                        monto_recoleccion_b = lista_limpia[posicion_10+2]
                        monto_recoleccion = monto_recoleccion_b.replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_recoleccion = ''
                        monto_recoleccion = ''

                    #Posicion 12 y 13
                    elemento_a_buscar = 'TRATAMIENTO'
                    try:
                        posicion_11 = lista_limpia.index(elemento_a_buscar)
                        cantidad_tratamiento = lista_limpia[posicion_11+1]
                        monto_tratamiento_b = lista_limpia[posicion_11+2]
                        monto_tratamiento = monto_tratamiento_b.replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_tratamiento = ''
                        monto_tratamiento = ''
                        
                    #Posicion 14
                    elemento_a_buscar = 'NETO'
                    try:
                        posicion_14 = lista_limpia.index(elemento_a_buscar)
                        neto_b = lista_limpia[posicion_14+1]
                        neto = neto_b.replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        neto = ''

                    #Posicion 15
                    elemento_a_buscar = 'IVA (19%)'
                    try:
                        posicion_15 = lista_limpia.index(elemento_a_buscar)
                        iva_b = lista_limpia[posicion_15+1]
                        iva = iva_b.replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        iva = ''

                    #Posicion 16
                    elemento_a_buscar = 'TOTAL VENTA'
                    try:
                        posicion_16 = lista_limpia.index(elemento_a_buscar)
                        total_venta_bru = lista_limpia[posicion_16+1]
                        total_venta = total_venta_bru.replace('.','') 
                    except:
                        print('elemento no se encuentra disponible')
                        total_venta = ''
                    
                    #Posicion 17
                    elemento_a_buscar = 'DESCUENTO LEY REDONDEO'
                    try:
                        posicion_17 = lista_limpia.index(elemento_a_buscar)
                        descuento_redondeo = lista_limpia[posicion_17+1]
                    except:
                        print('elemento no se encuentra disponible')
                        descuento_redondeo = ''
                    
                    #Posicion 18
                    elemento_a_buscar = 'TOTAL A PAGAR'
                    try:
                        posicion_17 = lista_limpia.index(elemento_a_buscar)
                        total_a_pagar_bruto = lista_limpia[posicion_17+1]
                        total_a_pagar = total_a_pagar_bruto.replace('$ ','').replace('.','') 
                    except:
                        print('elemento no se encuentra disponible')
                        descuento_redondeo = ''
                    

                    #Posicion 19 y 20
                    texto_a_verificar = 'LECTURA ACTUAL'
                    posicion_19 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_19 = idx
                            lectura_actual_bruto = lista_limpia[posicion_19]
                            lectura_actual_b = lectura_actual_bruto.replace('LECTURA ACTUAL ','')
                            partes = lectura_actual_b.split('-')
                            dia_lectura_actual = partes[0]
                            mes_lectura_actual = self.dic_datos(partes[1])
                            año_lectura_actual = partes[2]
                            fecha_lectura_actual = dia_lectura_actual+'-'+mes_lectura_actual+'-'+año_lectura_actual
                            valor_lectu_actual_b = lista_limpia[posicion_19+1]
                            valor_lectu_actual = valor_lectu_actual_b.replace(' m3','').replace('-','')
                            
                    #Posicion 21 y 22
                    texto_a_verificar = 'LECTURA ANTERIOR'
                    posicion_21 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_21 = idx
                            lectura_anterior_bruto = lista_limpia[posicion_21]
                            lectura_anterior_b = lectura_anterior_bruto.replace('LECTURA ANTERIOR ','')
                            partes = lectura_anterior_b.split('-')
                            dia_lectura_anterior = partes[0]
                            mes_lectura_anterior = self.dic_datos(partes[1])
                            año_lectura_anterior= partes[2]
                            fecha_lectura_anterior = dia_lectura_anterior+'-'+mes_lectura_anterior+'-'+año_lectura_anterior
                            valor_lectu_anterior_b = lista_limpia[posicion_21+1]
                            valor_lectu_anterior = valor_lectu_anterior_b.replace(' m3','').replace('-','')

                    #Posicion 23
                    elemento_a_buscar = 'DIFERENCIA DE LECTURAS'
                    try:
                        posicion_23 = lista_limpia.index(elemento_a_buscar)
                        diferencia_lecturas_b = lista_limpia[posicion_23+1]
                        diferencia_lecturas = diferencia_lecturas_b.replace(' m3','')
                    except:
                        print('elemento no se encuentra disponible')
                        diferencia_lecturas = ''
                        
                    #Posicion 24
                    elemento_a_buscar = 'CONSUMO PROMEDIO DESCONTABLE'
                    try:
                        posicion_24 = lista_limpia.index(elemento_a_buscar)
                        consum_promedio_desc_b = lista_limpia[posicion_24+1]
                        consum_promedio_desc = consum_promedio_desc_b.replace(' m3','')
                    except:
                        print('elemento no se encuentra disponible')
                        consum_promedio_desc = ''
                    
                    #Posicion 25
                    elemento_a_buscar = 'CONSUMO TOTAL'
                    try:
                        posicion_25 = lista_limpia.index(elemento_a_buscar)
                        consumo_total_b = lista_limpia[posicion_25+1]
                        consumo_total = consumo_total_b.replace(' m3','')
                    except:
                        print('elemento no se encuentra disponible')
                        consumo_total = ''
        
                    #Posicion 26
                    elemento_a_buscar = 'FECHA ESTIMADA PRÓXIMA LECTURA'
                    try:
                        posicion_26 = lista_limpia.index(elemento_a_buscar)
                        fecha_prox_lectura_b = lista_limpia[posicion_26+1]
                        partes = fecha_prox_lectura_b.split('-')
                        dia_prox_lectura = partes[0]
                        mes_prox_lectura = self.dic_datos(partes[1])
                        año_prox_lectura = partes[2]
                        fecha_proxima_lectura = dia_prox_lectura+'-'+mes_prox_lectura+'-'+año_prox_lectura
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_proxima_lectura = ''

                    #Posicion 27
                    elemento_a_buscar = 'Factor de Cobro del Periodo'
                    try:
                        posicion_27 = lista_limpia.index(elemento_a_buscar)
                        factor_cobro = lista_limpia[posicion_27+1]
                    except:
                        print('elemento no se encuentra disponible')
                        factor_cobro = ''

                    #Posicion 28
                    elemento_a_buscar = 'Punto Servicio-Diametro'
                    try:
                        posicion_28 = lista_limpia.index(elemento_a_buscar)
                        punto_ser_diametro = lista_limpia[posicion_28+1]
                    except:
                        print('elemento no se encuentra disponible')
                        punto_ser_diametro = ''

                    #Posicion 29
                    elemento_a_buscar = 'Clave Facturación'
                    try:
                        posicion_29 = lista_limpia.index(elemento_a_buscar)
                        clave_facturacion = lista_limpia[posicion_29+1]
                    except:
                        print('elemento no se encuentra disponible')
                        clave_facturacion = ''
                 
                    #Posicion 30
                    elemento_a_buscar = 'Clave Lectura'
                    try:
                        posicion_30 = lista_limpia.index(elemento_a_buscar)
                        clave_lectura = lista_limpia[posicion_30+1]
                    except:
                        print('elemento no se encuentra disponible')
                        clave_lectura = ''

                    #Posicion 31
                    texto_a_verificar = 'Metro cúbico agua potable punta'
                    posicion_31 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_31 = idx
                            mcubico_punta_br = lista_limpia[posicion_31]
                            mcubico_punta_b = mcubico_punta_br.split('= ')
                            mcubico_punta_ = mcubico_punta_b[1]
                            mcubico_punta = mcubico_punta_.replace('$ ','')
                        
                    #Posicion 32
                    texto_a_verificar = 'Metro cúbico agua potable no punta'
                    posicion_32 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_32 = idx
                            mcubico_no_punta_br = lista_limpia[posicion_32]
                            mcubico_no_punta_b = mcubico_no_punta_br.split('= ')
                            mcubico_no_punta_ = mcubico_no_punta_b[1]
                            mcubico_no_punta = mcubico_no_punta_.replace('$ ','')

                    #Posicion 33
                    texto_a_verificar = 'Metro cúbico sobreconsumo'
                    posicion_33 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_33 = idx
                            mcubico_sobreconsumo_br = lista_limpia[posicion_33]
                            mcubico_sobreconsumo_b = mcubico_sobreconsumo_br.split('= ')
                            mcubico_sobreconsumo_ = mcubico_sobreconsumo_b[1]
                            mcubico_sobreconsumo = mcubico_sobreconsumo_.replace('$ ','')

                    #Posicion 34
                    texto_a_verificar = 'Metro cúbico recolección'
                    posicion_34 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_34 = idx
                            mcubico_recoleccion_br = lista_limpia[posicion_34]
                            mcubico_recoleccion_b = mcubico_recoleccion_br.split('= ')
                            mcubico_recoleccion_ = mcubico_recoleccion_b[1]
                            mcubico_recoleccion = mcubico_recoleccion_.replace('$ ','')

                    #Posicion 35
                    texto_a_verificar = 'Metro cúbico tratamiento'
                    posicion_35 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_35 = idx
                            mcubico_tratamiento_br = lista_limpia[posicion_35]
                            mcubico_tratamiento_b = mcubico_tratamiento_br.split('= ')
                            mcubico_tratamiento_ = mcubico_tratamiento_b[1]
                            mcubico_tratamiento = mcubico_tratamiento_.replace('$ ','')

                    #Posicion 36
                    texto_a_verificar = 'Corte o Reposición 1era instancia'
                    posicion_36 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_36 = idx
                            primera_instancia_br = lista_limpia[posicion_36]
                            primera_instancia_b = primera_instancia_br.split(': ')
                            primera_instancia_ = primera_instancia_b[1]
                            primera_instancia = primera_instancia_.replace('$ ','').replace('.','')

                    #Posicion 37
                    texto_a_verificar = 'Corte o Reposición 2da instancia'
                    posicion_37 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_37 = idx
                            segunda_instancia_br = lista_limpia[posicion_37]
                            segunda_instancia_b = segunda_instancia_br.split(': ')
                            segunda_instancia_ = segunda_instancia_b[1]
                            segunda_instancia = segunda_instancia_.replace('$ ','').replace('.','')

                    #Posicion 38
                    elemento_a_buscar = 'Número de Medidor'
                    try:
                        posicion_38 = lista_limpia.index(elemento_a_buscar)
                        n_medidor = lista_limpia[posicion_38+1]
                    except:
                        print('elemento no se encuentra disponible')
                        n_medidor = ''  
                        
                    #Posicion 39
                    elemento_a_buscar = 'Grupo Tarifario'
                    try:
                        posicion_39 = lista_limpia.index(elemento_a_buscar)
                        grupo_tarifario = lista_limpia[posicion_39+1]
                    except:
                        print('elemento no se encuentra disponible')
                        grupo_tarifario = ''

                    #Posicion 40
                    texto_a_verificar = 'FECHA EMISIÓN'
                    posicion_40 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_40 = idx
                            fecha_emision_br = lista_limpia[posicion_40]
                            fecha_emision_b = fecha_emision_br.split(':')
                            fecha_emision_ = fecha_emision_b[1]
                            partes = fecha_emision_.split('-')
                            dia_emision = partes[0]
                            mes_emision = self.dic_datos(partes[1])
                            año_emision = partes[2]
                            fecha_emision = dia_emision+'-'+mes_emision+'-'+año_emision
                    
                #Cargamos libro excel donde volcaremos los datos
                libro = load_workbook(output_path+'/'+'Formato Planilla.xlsx')
                hoja_agua = libro['Agua']
                    
                ultima_fila = hoja_agua.max_row
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=3).value = int(n_factura)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=3).value = n_factura  
                    
                hoja_agua.cell(row=ultima_fila+1,column=8).value = rut
                
                hoja_agua.cell(row=ultima_fila+1,column=9).value = n_cuenta
                hoja_agua.cell(row=ultima_fila+1,column=7).value = giro
                hoja_agua.cell(row=ultima_fila+1,column=11).value = fecha_vencimiento
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=46).value = int(total_a_pagar)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=46).value = total_a_pagar
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=14).value = int(cargo_fijo)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=14).value = cargo_fijo
                    
                hoja_agua.cell(row=ultima_fila+1,column=15).value = cantidad_consumo_agua
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=17).value = int(monto_consumo_agua)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=17).value = monto_consumo_agua
                    
                hoja_agua.cell(row=ultima_fila+1,column=72).value = cantidad_recoleccion
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=74).value = int(monto_recoleccion)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=74).value = monto_recoleccion
                
                hoja_agua.cell(row=ultima_fila+1,column=28).value = cantidad_tratamiento
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=30).value = int(monto_tratamiento)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=30).value = monto_tratamiento
                
                try:  
                    hoja_agua.cell(row=ultima_fila+1,column=42).value = int(neto)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=42).value = neto
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=43).value = int(iva)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=43).value = iva
                
                try:  
                    hoja_agua.cell(row=ultima_fila+1,column=44).value = int(total_venta)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=44).value = total_venta
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=46).value = int(total_a_pagar)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=46).value = total_a_pagar
                
                hoja_agua.cell(row=ultima_fila+1,column=57).value = fecha_lectura_actual
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=56).value = int(valor_lectu_actual)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=56).value = valor_lectu_actual
                    
                hoja_agua.cell(row=ultima_fila+1,column=58).value = fecha_lectura_anterior
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=76).value = int(valor_lectu_anterior)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=76).value = valor_lectu_anterior
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=77).value = int(diferencia_lecturas)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=77).value = diferencia_lecturas
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=66).value = int(consumo_total)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=66).value = consumo_total
                
                hoja_agua.cell(row=ultima_fila+1,column=54).value = fecha_proxima_lectura
                hoja_agua.cell(row=ultima_fila+1,column=61).value = factor_cobro
                hoja_agua.cell(row=ultima_fila+1,column=53).value = punto_ser_diametro
                hoja_agua.cell(row=ultima_fila+1,column=59).value = clave_lectura
                
                hoja_agua.cell(row=ultima_fila+1,column=42).value = mcubico_punta
                hoja_agua.cell(row=ultima_fila+1,column=36).value = mcubico_no_punta
                hoja_agua.cell(row=ultima_fila+1,column=39).value = mcubico_sobreconsumo
                hoja_agua.cell(row=ultima_fila+1,column=73).value = mcubico_recoleccion
                hoja_agua.cell(row=ultima_fila+1,column=29).value = mcubico_tratamiento
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=78).value = int(primera_instancia)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=78).value = primera_instancia
                
                try:   
                    hoja_agua.cell(row=ultima_fila+1,column=79).value = int(segunda_instancia)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=79).value = segunda_instancia
                
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=52).value = int(n_medidor)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=52).value = n_medidor
                    
                hoja_agua.cell(row=ultima_fila+1,column=50).value = grupo_tarifario
                hoja_agua.cell(row=ultima_fila+1,column=10).value = fecha_emision
                
                libro.save(output_path+'/'+'Formato Planilla.xlsx')
                    
                #Copiamos el archivo a la carpeta outpu con el nombre que corresponde
                shutil.copy(archivo, output_path+nombre_oficial)
                print('-----')
    
        #Obtenemos los archivos de la carpeta input
        archivos_en_carpeta = os.listdir(folder_path)

        # Iterar sobre los archivos y eliminarlos
        for archivo in archivos_en_carpeta:
            ruta_archivo = os.path.join(folder_path, archivo)
            if os.path.isfile(ruta_archivo):
                os.remove(ruta_archivo)