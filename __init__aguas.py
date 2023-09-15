from code.app_aguas import Scraper_Aguas
#import smtplib
from openpyxl import load_workbook

def send_notification():
    # Código para enviar correo electrónico de notificación
    print('')

if __name__ == '__main__':

    print('Obteniendo credenciales...')
    print('----------------------------------------------------------------------')
        
    credencials = '.\config\credenciales.xlsx'
    libro_accesos = load_workbook(credencials)
    hoja_credenciales = libro_accesos['Hoja1']
        
    for j in hoja_credenciales.iter_rows(2):
        try:
            rut = j[0].value
            passw = j[1].value
            web = j[2].value
            break
        except:
            ('no hay credenciales')
            
    email = rut
    password = passw
    url = web
    driver_path = 'chromedriver.exe'
    
    scraper = Scraper_Aguas(url, email, password, driver_path)
    #Primera sociedad
    print('ingresamos en la clase Scraper_Aguas...')
    print('----------------------------------------------------------------------')
    
    scraper.login()
    print('hacemos login en el portal...')
    print('----------------------------------------------------------------------')
    
    scraper.scrapping_aguas(posicion=0)
    print('hacemos scrapping al portal 1era sociedad...')
    print('----------------------------------------------------------------------')

    scraper.close()
    print('cerramos el bot 1era parte...')
    print('----------------------------------------------------------------------')
    
    #Segunda sociedad
    print('segunda sociendad...')
    print('----------------------------------------------------------------------')
    
    scraper.login()
    print('hacemos login nuevamente en el portal...')
    print('----------------------------------------------------------------------')
    
    scraper.scrapping_aguas(posicion=1)
    print('hacemos scrapping al portal 2da. sociedad...')
    print('----------------------------------------------------------------------')
    
    scraper.archivos()
    print('procesando datos hacia Planilla Formato')
    print('----------------------------------------------------------------------')
    
    scraper.close()
    print('cerramos el bot final...')
    print('----------------------------------------------------------------------')
